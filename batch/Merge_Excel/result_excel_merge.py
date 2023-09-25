import openpyxl
from openpyxl import styles
import os
import tkinter as tk
import tkinter.messagebox
from openpyxl.utils import get_column_letter

from self_def_functions import copy_data

input_dir = r'.\sample_files'
output_path = r'.\output\sum_py.xlsx'

#切换工作路径至py文件路径
py_dir = os.path.dirname(__file__)
os.chdir(py_dir)

if(os.path.exists(output_path)):
    tkinter.messagebox.showwarning(title='提示', message='文件已存在')

else:
    #新建Excel，Sheet名为“汇总”
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '汇总'

    #读取输入文件夹内的文件列表，并处理为相对路径
    file_list = os.listdir(input_dir)
    file_list_full = []

    for item in file_list:
        file_list_full.append(os.path.join(input_dir, item))

    wb_header = openpyxl.load_workbook(file_list_full[0])
    ws_list = wb_header.sheetnames
    ws_header = wb_header[ws_list[0]]
    data_header = ws_header['A1':'R2']
    dest_cell = ws_out['A1']
    # print(ws_header.max_row)
            
    copy_data(data_header, dest_cell)

    for file in file_list_full:
        wb_src = openpyxl.load_workbook(file)
        ws_src_list = wb_src.sheetnames
        ws_src = wb_src[ws_src_list[0]]
        maxrow_src = str(ws_src.max_row)
        data = ws_src['A3':('R'+maxrow_src)]

        maxrow_out = str(ws_out.max_row)
        dest_cell_out = ws_out['A'+maxrow_out]
        copy_data(data, dest_cell_out)

    #设置表头样式
    header_style = styles.NamedStyle(name='header_style')
    header_style.font = styles.Font(bold=True)
    base_border = styles.borders.Side(border_style='thin')
    header_style.border = styles.Border(left=base_border,right=base_border,top=base_border,bottom=base_border)
    header_style.alignment = styles.Alignment(vertical='center', horizontal='center')
    header_style.fill = styles.PatternFill(fill_type='gray125', start_color='CCCCCC', end_color='000000')

    predef_width = [5.13, 10.38, 10.38, 6.5, 16.63, 7.88, 6.5, 15.5, 60.00, 14.88, 16.00, 27.00, 14.5, 12.38, 14.5, 19.8, 7.13, 13]
    for i in range(1, ws_out.max_column, 1):
        ws_out.cell(row=1, column=i).style = header_style
        ws_out.cell(row=2, column=i).style = header_style
        # print('col '+get_column_letter(i)+' width is '+str(predef_width[i-1]))
        ws_out.column_dimensions[get_column_letter(i)].width = predef_width[i-1]

    merge_row = ['A', 'B', 'P', 'Q', 'R']
    for col in merge_row:
        ws_out.merge_cells(col+'1:'+col+'2')
    ws_out.merge_cells('C1:I1')
    ws_out.merge_cells('J1:O1')
    # ws_out['A1':'R2'].border = border
    # ws_out.page_setup.fitToWidth = 1
    wb_out.save(output_path)