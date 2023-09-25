import os
import difflib
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl import styles
import tkinter.messagebox as messagebox
import tkinter as tk

input_file = r'./(合同编号CMITYD-202200099)-付款凭证-基础平台部-领航动力信息系统有限公司2023Q3（研发）.xlsx'
summary_log = os.path.splitext(input_file)[-2]+'_check_log.txt'

similar_threshold = 0.95

#切换工作路径至py文件路径
py_dir = os.path.dirname(__file__)
os.chdir(py_dir)

def output(idx_list: list, description_str: str, success_str: str, column_name: str, color: str, reference_list = []) -> str:
    """
    输出函数，将idx_list中的每一项输出在description_str之后，idx_list为空则输出success_str，column_name为列名，用于表示输出的单元格定位。
    """
    if(len(idx_list)):
        output = description_str +'\n'
        for i in idx_list:
            output += column_name+str(i)+' '
            if(i in reference_list):
                ws[column_name+str(i)].fill = styles.PatternFill(fill_type='solid', start_color='6fe35a', end_color='000000')
            else:
                ws[column_name+str(i)].fill = styles.PatternFill(fill_type='solid', start_color=color, end_color='000000')
    else: output = success_str
    return output

def get_real_maxrow(ws) ->list:
    """
    定位需要进行判断的部分的位置，利用
    """
    row_num = ws.max_row
    for j in range(1,row_num):
        if(isinstance(ws.cell(row=j, column=1).value, int)): 
            break
    for i in range(j,row_num):
        if(not isinstance(ws.cell(row=i, column=1).value, int)): 
            break
    return [j, i]

#读Excel的第一张表
wb = openpyxl.load_workbook(input_file)
ws = wb['付款凭证-工作量核算表']
row_num = ws.max_row
row_num = get_real_maxrow(ws)

false_workload_num_idx = []
false_name_idx = []
empty_name_idx = []
false_description_idx = []
similar_name_idx = []
similar_name_idx_set = set()

for i in range(get_real_maxrow(ws)[0], get_real_maxrow(ws)[1]):
    # 判断工作量
    cell_value = ws.cell(row=i, column=column_index_from_string('I')).value
    if(int(cell_value) < 0 or int(cell_value) >= 50):
        false_workload_num_idx.append(i)
    
    # 查找关键字
    cell_value = ws.cell(row=i, column=column_index_from_string('E')).value
    flag = 0
    for keyword in ['设计','调研','专利','软著']:
        if(cell_value != None):
            if keyword in cell_value: flag = flag + 1
    if(flag):
        false_name_idx.append(i)

    flag = 0
    for keyword in ['运营','运维']:
        if(cell_value != None):
            if keyword in cell_value: flag = flag + 1
    if(flag):
        false_name_idx.append(i)

    cell_value = ws.cell(row=i, column=column_index_from_string('F')).value
    flag = 0
    for keyword in ['运营','运维','设计','调研','专利','软著']:
        if keyword in cell_value: flag = flag + 1
    if(flag):
        false_description_idx.append(i)

    # 判断需求名称是否有空
    cell_value = ws.cell(row=i, column=column_index_from_string('E')).value
    if(cell_value == None):
        empty_name_idx.append(i)

# 检查相似
for i in range(get_real_maxrow(ws)[0], get_real_maxrow(ws)[1]):
    cell_value = ws.cell(row=i, column=column_index_from_string('E')).value
    if(cell_value == None): cell_value = ''
    for j in range(i+1,get_real_maxrow(ws)[1]):
        compare_cell_value = ws.cell(row=j, column=column_index_from_string('E')).value
        if(compare_cell_value == None): compare_cell_value = ''
        if(difflib.SequenceMatcher(None, cell_value, compare_cell_value).quick_ratio() > similar_threshold):
            similar_name_idx.append([i,j])
            similar_name_idx_set.update([i,j])
similar_name_idx_set = sorted(similar_name_idx_set)

result = (#输出工作量超范围的单元格位置
        output(false_workload_num_idx, '工作量超出范围 (0<X<50):', '工作量符合要求', 'I', color='fff94a')+'\n\n'+
        #输出需求名包含关键字的单元格位置
        output(false_name_idx, '需求名称中包含关键字 (运营、运维、设计、调研、专利、软著):', '需求名称符合要求', 'E', color='fff94a')+'\n\n'+
        #输出需求描述包含关键字的单元格位置
        output(false_description_idx, '详细描述中包含关键字 (运营、运维、设计、调研、专利、软著):', '详细描述符合要求', 'F', color='fff94a')+'\n\n'+
        #输出需求名为空的单元格位置
        output(empty_name_idx, '三级需求为空的单元格:', '三级需求均已填写', 'E', color='fff94a')+'\n\n'+
        #输出相似需求名
        output(similar_name_idx_set, '疑似类似的需求名称:', '无疑似类似的需求名称', 'E', color='5abde3', reference_list=false_name_idx)+'\n\n')

with open(summary_log, 'w', encoding='utf8') as f:
    f.write("文件 <"+input_file+"> 的检查报告\ncheck.xlsx需求名称列中，黄色框为关键字有问题，蓝色框为相似，绿色框为同时涉及以上两项\n--------------------------------\n\n")
    f.write(result)
    f.write(
        '以下行中需求名称之间较为类似：\n'+
        str(similar_name_idx)
    )

copy_file = os.path.splitext(input_file)[-2]+'_check'+os.path.splitext(input_file)[-1]
wb.save(copy_file)

root = tk.Tk()
root.withdraw()

messagebox.showinfo(title='提示', message=result+'日志保存为'+summary_log)

# print(45.1> 0 and 45.1< 50)
# similarity = difflib.SequenceMatcher(None, str1, str2).quick_ratio()
